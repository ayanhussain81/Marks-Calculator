


**************Created By Ayan Hussain**************

#Importing the liberaries

from tkinter import *
from openpyxl import load_workbook

m = Tk()
 
        #====================Variables========================#

ds1 = DoubleVar()
pl1 = DoubleVar()
cal1 = DoubleVar()
eng1 = DoubleVar()
fit1 = DoubleVar()
pst1 = DoubleVar()
rno = IntVar()
naam = StringVar()
fnaam = StringVar()
dob = StringVar()
grp = StringVar()
gen = StringVar()

#Funtion to sum the marks 
def add():
    g = ds1.get()
    b = pl1.get()
    c = cal1.get()
    d = eng1.get()
    e = pst1.get()
    f = fit1.get()
    s = g + b + c + d + e + f
    l99.config(text=s, fg="black")
    return s

#Function to calculate the percentage 
def percent():
    s = add()
    z = (s / 600) * 100
    e26.config(text=z, fg="black")
    return z

#Function to calculate the grade
def grade():
    list = []
    z = percent()
    if z >= 80:
        e27.config(text="A+", fg="green")
        list.append("A+")
    elif z < 80 and z >= 70:
        e27.config(text="A", fg="green")
        list.append("A")

    elif z < 70 and z >= 60:
        e27.config(text="B", fg="green")
        list.append("B")
    elif z < 60 and z >= 50:
        e27.config(text="C", fg="green")
        list.append("C")
    elif z < 50 and z >= 40:
        e27.config(text="D", fg="green")
        list.append("D")
    else:
        e27.config(text="FAIL", fg="Red")
        list.append("FAIL")
    return list

         #====================Button Funstions========================#

def combine():
    add()
    percent()
    grade()


def submit():
    s = add()
    z = percent()
    list = grade()
    wb = load_workbook('C:/Users/92336/Desktop/wb.xlsx')    # Path of the Excel file
    ws = wb.active
    current_row = ws.max_row
    # current_column = ws.max_column
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 17
    ws.column_dimensions['I'].width = 13
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 25
    ws.column_dimensions['M'].width = 25
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 15
    ws.cell(row=1,column=1).value="ROLL NO"
    ws.cell(row=1, column=2).value ="NAME"
    ws.cell(row=1, column=3).value="FATHER NAME"
    ws.cell(row=1, column=4).value="DATE OF BIRTH"
    ws.cell(row=1, column=5).value="GROUP"
    ws.cell(row=1, column=6).value="GENDER"
    ws.cell(row=1, column=7).value="DISCRETE STRUCTURE"
    ws.cell(row=1, column=8).value="PROGRAMMING"
    ws.cell(row=1, column=9).value="CALCULUS"
    ws.cell(row=1, column=10).value="ENGLISH"
    ws.cell(row=1, column=11).value="PAKISTAN STUDIES"
    ws.cell(row=1, column=12).value="FUNDAMENTALS OF IT"
    ws.cell(row=1, column=13).value="TOTAL OBTAINED MARKS"
    ws.cell(row=1, column=14).value="PERCENTAGE"
    ws.cell(row=1, column=15).value ="GRADE"
    ws.cell(row=current_row+1,column=1).value=int(rno.get())
    ws.cell(row=current_row + 1, column=2).value = str(naam.get())
    ws.cell(row=current_row + 1, column=3).value=str(fnaam.get())
    ws.cell(row=current_row + 1, column=4).value=str(dob.get())
    ws.cell(row=current_row + 1, column=5).value=str(grp.get())
    ws.cell(row=current_row + 1, column=6).value=str(gen.get())
    ws.cell(row=current_row + 1, column=7).value=int(ds1.get())
    ws.cell(row=current_row + 1, column=8).value=int(pl1.get())
    ws.cell(row=current_row + 1, column=9).value=int(cal1.get())
    ws.cell(row=current_row + 1, column=10).value=int(eng1.get())
    ws.cell(row=current_row + 1, column=11).value=int(pst1.get())
    ws.cell(row=current_row + 1, column=12).value=int(fit1.get())
    ws.cell(row=current_row + 1, column=13).value=int(s)
    ws.cell(row=current_row + 1, column=14).value=int(z)
    for j in list:
        ws.cell(row=current_row + 1, column=15).value=str(j)
    wb.save('C:/Users/92336/Desktop/wb.xlsx')



def delete():
    e0.delete(0, END)
    e1.delete(0, END)
    e2.delete(0, END)
    e3.delete(0, END)
    e5.delete(0, END)
    ee5.delete(0, END)
    ds.delete(0, END)
    pl.delete(0, END)
    calculus.delete(0, END)
    english.delete(0, END)
    pst.delete(0, END)
    fit.delete(0,END)


# m.geometry("700x800")
a = Label(m, text="STUDENT MARKSHEET", bg="seashell3",fg="navy", font=("Arial Black", "40"),relief=GROOVE,borderwidth=5)
a.pack(side=TOP, fill=X)
f1 = Frame(m, bg="Peach Puff2",relief=GROOVE,borderwidth=5)
f1.place(x=5, y=85, width=550, height=550)

l0 = Label(f1, text="ROLL NO.", width=20, bg="Peach Puff2", font=("Times New Roman", 10), relief=SUNKEN)
l0.grid(row=0, column=0, padx=10, pady=20)
e0 = Entry(f1, bg="azure", textvariable=rno)
e0.grid(row=0, column=1, padx=70, pady=20)

l1 = Label(f1, text="NAME", width=20, bg="Peach Puff2", font=("Times New Roman", 10), relief=SUNKEN)
l1.grid(row=1, column=0, padx=10, pady=20)
e1 = Entry(f1, bg="azure", textvariable=naam)
e1.grid(row=1, column=1, padx=70, pady=20)

l2 = Label(f1, text="FATHER NAME", width=20, bg="Peach Puff2", font=("Times New Roman", 10), relief=SUNKEN)
l2.grid(row=2, column=0, padx=10, pady=20)
e2 = Entry(f1, bg="azure", textvariable=fnaam)
e2.grid(row=2, column=1, padx=70, pady=20)

l3 = Label(f1, text="DATE OF BIRTH", width=20, bg="Peach Puff2", font=("Times New Roman", 10), relief=SUNKEN)
l3.grid(row=3, column=0, padx=10, pady=20)
e3 = Entry(f1, bg="azure", textvariable=dob)
e3.grid(row=3, column=1, padx=70, pady=20)

l5 = Label(f1, text="GROUP", width=20, bg="Peach Puff2", font=("Times New Roman", 10), relief=SUNKEN)
l5.grid(row=4, column=0, padx=10, pady=20)
e5 = Entry(f1, bg="azure", textvariable=grp)
e5.grid(row=4, column=1, padx=70, pady=20)

l6 = Label(f1, text="GENDER", width=20, bg="Peach Puff2", font=("Times New Roman", 10), relief=SUNKEN)
l6.grid(row=5, column=0, padx=10, pady=20)
ee5 = Entry(f1, bg="azure", textvariable=gen)
ee5.grid(row=5, column=1, padx=70, pady=20)

# male = Radiobutton(f1,text='Male',variable="btn1",value='male',bg="ivory3",width=10)
# male.grid(row=5,column=1)
# female = Radiobutton(f1,text='Female',variable="btn2",value='female',bg="ivory3",width=10)
# female.grid(row=6,column=1)

f2 = Frame(m, bg="light cyan",relief=GROOVE,borderwidth=5)
f2.place(x=550, y=85, width=1200, height=550)

l7 = Label(f2, text="COURSES", bg="Light cyan", width=20, font=("Times New Roman", 12), relief=SUNKEN)
l7.grid(row=0, column=0, padx=15, pady=20)

l8 = Label(f2, text="MARKS OBTAINED", bg="Light cyan", width=20, font=("Times New Roman", 12), relief=SUNKEN)
l8.grid(row=0, column=1, padx=15, pady=20)

l9 = Label(f2, text="PASSING MARKS", bg="Light cyan", width=20, font=("Times New Roman", 12), relief=SUNKEN)
l9.grid(row=0, column=2, padx=15, pady=20)

l10 = Label(f2, text="TOTAL MARKS", bg="Light cyan", width=20, font=("Times New Roman", 12), relief=SUNKEN)
l10.grid(row=0, column=3, padx=15, pady=20)

l11 = Label(f2, text="DISCRETE STRUCTURE", bg="Light cyan", width=20, font=("Arial", 10), relief=GROOVE)
l11.grid(row=1, column=0, padx=15, pady=15)

l12 = Label(f2, text="PROGRAMING", bg="Light cyan", width=20, font=("Arial", 10), relief=GROOVE)
l12.grid(row=2, column=0, padx=15, pady=15)

l13 = Label(f2, text="CALCULUS", bg="Light cyan", width=20, font=("Arial", 10), relief=GROOVE)
l13.grid(row=3, column=0, padx=15, pady=15)

l14 = Label(f2, text="ENGLISH", bg="Light cyan", width=20, font=("Arial", 10), relief=GROOVE)
l14.grid(row=4, column=0, padx=15, pady=15)

l15 = Label(f2, text="PAKISTAN STUDIES", bg="Light cyan", width=20, font=("Arial", 10), relief=GROOVE)
l15.grid(row=5, column=0, padx=15, pady=15)

l12 = Label(f2, text="FUNDAMENTALS OF IT", bg="Light cyan", width=20, font=("Arial", 10), relief=GROOVE)
l12.grid(row=6, column=0, padx=15, pady=15)

ds = Entry(f2, bg="azure", width=15, textvariable=ds1, font=("Arial Black", 10), relief=GROOVE)
ds.grid(row=1, column=1,padx=15, pady=15)
pl = Entry(f2, bg="azure", width=15, textvariable=pl1, font=("Arial Black", 10), relief=GROOVE)
pl.grid(row=2, column=1,padx=15, pady=15)
calculus = Entry(f2, bg="azure", width=15, textvariable=cal1, font=("Arial Black", 10), relief=GROOVE)
calculus.grid(row=3,column=1,padx=15,pady=15)
english = Entry(f2, bg="azure", width=15, textvariable=eng1, font=("Arial Black", 10), relief=GROOVE)
english.grid(row=4,column=1,padx=15,pady=15)
pst = Entry(f2, bg="azure", width=15, textvariable=pst1, font=("Arial Black", 10), relief=GROOVE)
pst.grid(row=5, column=1,padx=15, pady=15)
fit = Entry(f2, bg="azure", width=15, textvariable=fit1, font=("Arial Black", 10), relief=GROOVE)
fit.grid(row=6, column=1,padx=15, pady=15)

e12 = Label(f2, bg="azure", width=15, text=33, font=("Arial Black", 10), relief=GROOVE).grid(row=1, column=2, padx=15,
                                                                                             pady=15)
e13 = Label(f2, bg="azure", width=15, text=33, font=("Arial Black", 10), relief=GROOVE).grid(row=2, column=2, padx=15,
                                                                                             pady=15)
e14 = Label(f2, bg="azure", width=15, text=33, font=("Arial Black", 10), relief=GROOVE).grid(row=3, column=2, padx=15,
                                                                                             pady=15)
e15 = Label(f2, bg="azure", width=15, text=33, font=("Arial Black", 10), relief=GROOVE).grid(row=4, column=2, padx=15,
                                                                                             pady=15)
e16 = Label(f2, bg="azure", width=15, text=33, font=("Arial Black", 10), relief=GROOVE).grid(row=5, column=2, padx=15,
                                                                                             pady=15)
e17 = Label(f2, bg="azure", width=15, text=33, font=("Arial Black", 10), relief=GROOVE).grid(row=6, column=2, padx=15,
                                                                                             pady=15)

e18 = Label(f2, bg="azure", width=15, text=100, font=("Arial Black", 10), relief=GROOVE).grid(row=1, column=3, padx=15,
                                                                                              pady=15)
e19 = Label(f2, bg="azure", width=15, text=100, font=("Arial Black", 10), relief=GROOVE).grid(row=2, column=3, padx=15,
                                                                                              pady=15)
e20 = Label(f2, bg="azure", width=15, text=100, font=("Arial Black", 10), relief=GROOVE).grid(row=3, column=3, padx=15,
                                                                                              pady=15)
e21 = Label(f2, bg="azure", width=15, text=100, font=("Arial Black", 10), relief=GROOVE).grid(row=4, column=3, padx=15,
                                                                                              pady=15)
e22 = Label(f2, bg="azure", width=15, text=100, font=("Arial Black", 10), relief=GROOVE).grid(row=5, column=3, padx=15,
                                                                                              pady=15)
e23 = Label(f2, bg="azure", width=15, text=100, font=("Arial Black", 10), relief=GROOVE).grid(row=6, column=3, padx=15,
                                                                                              pady=15)

f3 = Frame(m, bg="Antique White2",relief=GROOVE,borderwidth=5)
f3.place(x=5, y=500, width=1600, height=600)
total_marks_obtained = Label(f3, text="TOTAL OBTAINED MARKS", bg="Antique White2", width=20, font=("Arial Black", 11),
                             relief=RAISED)
total_marks_obtained.grid(row=0, column=0, padx=20, pady=15)
l99 = Label(f3, bg="azure", text=0, width=20, fg="grey", font=("Arial Black", 11), relief=RAISED)
l99.grid(row=0, column=1, padx=20, pady=20, ipady=3)

total_marks = Label(f3, text="TOTAL MARKS", bg="Antique White2", width=20, font=("Arial Black", 11), relief=RAISED)
total_marks.grid(row=1, column=0, padx=20, pady=15)
e25 = Label(f3, bg="azure", width=20, text=600, font=("Arial Black", 11), relief=RAISED)
e25.grid(row=1, column=1, padx=20, pady=20, ipady=3)

percentage = Label(f3, text="PERCENTAGE", bg="Antique White2", width=20, font=("Arial Black", 11), relief=RAISED)
percentage.grid(row=0, column=3, padx=20, pady=15)
e26 = Label(f3, bg="azure", text=0, width=20, font=("Arial Black", 11), relief=RAISED)
e26.grid(row=0, column=4, padx=20, pady=20, ipady=3)

gra = Label(f3, text="GRADE", bg="Antique White2", width=20, font=("Arial Black", 11), relief=RAISED)
gra.grid(row=1, column=3, padx=20, pady=15)
e27 = Label(f3, bg="azure", width=20, font=("Arial Black", 11), relief=RAISED)
e27.grid(row=1, column=4, padx=20, pady=20, ipady=3)

b1 = Button(f3, text="COMPUTE", width=20, bg="Red", font=("Arial Black", 10), relief=GROOVE, command=combine)
b1.grid(row=0, column=5, padx=5)

b2 = Button(f3, text="SUBMIT", width=20, bg="Red", font=("Arial Black", 10), relief=GROOVE, command=submit)
b2.grid(row=1, column=5, padx=5)

b3 = Button(f3, text="CLEAR", width=20, bg="Red", font=("Arial Black", 10), relief=GROOVE, command=delete)
b3.grid(row=2, column=5, padx=5,pady=15)
m.mainloop()

**************Created By Ayan Hussain**************
